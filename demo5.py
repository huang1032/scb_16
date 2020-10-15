import requests
import openpyxl
import pprint


#读取excel表数据
def ReadData(filename,sheetlname):
    # 获取工作簿
    wk = openpyxl.open(filename=filename)
    # 获取当前工作表
    sheetl = wk[sheetlname]
    #print(sheetl.cell(row=1,column=1).value) #获取单元格1行1列的值
    # 获取数据表中所有数据的值
    #创建一个列表，把数据装起来
    datalist=[]
    for x in range(2,sheetl.max_row+1):
        # 把用例存储到字典
        case=dict(ids=sheetl.cell(x,1).value,
                 url=sheetl.cell(row=x, column=5).value,
                data=sheetl.cell(row=x, column=6).value,
            expected=sheetl.cell(row=x, column=7).value)
        #把每一个用例添加到列表中
        datalist.append(case)
        # 需要设置返回值
    return datalist

# 发送请求--封装成一个函数
def api_request(url,json_data):
    url=url       #发送一个url就传一个对应的url
    json=json_data
    # 设置成默认值，所有请求都是同一个请求头，请求头写死
    headers={'X-Lemonban-Media-Type':'lemonban.v2',
             "Content-Type":"application/json"}
    # 发送请求，传入url，json_data，headers
    response=requests.post(url=url,json=json_data,headers=headers)
    return response.json() #返回json格式的数据，后续模块需要需要用到这个数据则需要设置返回值

# 回写测试执行结果方法---openpyxl编辑excel文件
def write_result(filename,sheetname,row,column,result):
    wk1=openpyxl.open(filename=filename)
    register=wk1[sheetname]
    register.cell(row,column).value=result
    wk1.save(filename)

def excel_case(filename,sheetname):
    # 执行测试用例
    cases=ReadData(filename,sheetname)
    # 遍历列表中每一条用例
    for case in cases:
        id=case['ids']
        url=case['url']
        # 去引号把data数据转换为字典
        data=eval(case.get("data"))
        # print(case.get("data"),type(case.get("data")))
        '''---注释----------------------
        eval处理字符串表达式，返回表达式的值，它也可以把字符串转化为字典、元组、列表
        eval('3*2' ) 返回：6    eval('(1,2,3,4,5)') 返回(1,2,3,4,5)--元组tuple
        '''
        # 响应数据
        response=api_request(url=url,json_data=data)
        print(response)
        # 比对预期结果和实际结果
        expected_msg=eval(case['expected'])['msg']  # 获取expected期望结果中msg的值,并转换expected格式
        real_msg=response['msg']    # 获取expected实际响应结果中msg的值
        #进行比对
        if real_msg==expected_msg:
            print(f"用例{id}测试执行通过")
            write_result(filename=filename,sheetname=sheetname,
                         row=id+1,column=8,result='通过')
        else:
            print(f"用例{id}测试执行不通过")
            write_result(filename=filename, sheetname=sheetname,
                         row=id + 1, column=8, result='不通过')
        print('*'*30)    #输出30个*来进行分隔

# 调用
excel_case('test_case_api.xlsx','register')

# 登录
# excel_case('test_case_api.xlsx','login')